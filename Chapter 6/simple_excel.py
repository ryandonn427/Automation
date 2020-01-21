import openpyxl

xlsxfile = openpyxl.load_workbook('movies.xlsx')

print(xlsxfile.sheetnames)

sheet = xlsxfile['Sheet1']

print(sheet['B4'].value)
print(sheet['D4'].value)

print(sheet.max_row)
print(sheet.max_column)

print(sheet['A12'].value)
print(sheet['E1'].value)
