import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

xlsfile = openpyxl.Workbook()

data = [
    ('Name', 'Admissions'),
    ('Gone With the Wind', 225.7),
    ('Star Wars', 194.4),
    ('ET: The Extraterrestrial' ,161.0),
    ('The Sound of Music', 156.4)
]

sheet = xlsfile['Sheet']
for row in data:
    sheet.append(row)

BLUE = "0033CC"
LIGHT_BLUE = "E6ECFF"
WHITE = "FFFFFF"

header_font = Font(name='Tahoma',size=14,color=WHITE)
header_fill = PatternFill("solid",fgColor=BLUE)
for row in sheet['A1:B1']:
    for cell in row:
        cell.font = header_font
        cell.fill = header_fill

white_side = Side(border_style='thin', color = WHITE)
blue_side = Side(border_style='thin',color = BLUE)
alternate_fill = PatternFill("solid",fgColor=LIGHT_BLUE)
border = Border(bottom=blue_side, left=white_side,right=white_side)
for row_index, row in enumerate(sheet['A2:B5']):
    for cell in row:
        cell.border = border
        if row_index %2:
            cell.fill = alternate_fill

xlsfile.save('movies_format.xlsx')